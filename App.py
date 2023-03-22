import openpyxl as opx
import PySimpleGUI as sg

# Login
def sgLogin():
    layoutLogin = [
        [sg.Image(filename="./logo.png")],
        [sg.Text("Usuário"),],
        [sg.Input(key="usuario")],
        [sg.Text("Senha")],
        [sg.Input(key="senha", password_char="*")],
        [sg.Text("", key="mensagem")],
        [sg.Button("Login", button_color="#2f56d8", size=(10)), 
        sg.Button("Sair", button_color="#cc0000", size=(5))],
    ]
    
    janelaLogin = sg.Window("LOGIN", layout=layoutLogin, font="Roboto 15", size=(500,400))

    while True:
        event, values = janelaLogin.read()
        if event == "Sair":
            sg.Exit()
            break
        elif event == sg.WIN_CLOSED:
            break
        elif event == "Login":
            # ---Acesso ao sistema---
            usuarioCorreto = "admin"
            senhaCorreta = "1234"
            # -----------------------
            usuario = values["usuario"]
            senha = values["senha"]
            if senha == senhaCorreta and usuario == usuarioCorreto:
                janelaLogin["mensagem"].update("Acesso realizado com sucesso!", text_color="#00FF00")
                janelaLogin.close()
                sgMenu()
            else:
                janelaLogin["mensagem"].update("Dados de acesso incorreto(s).", text_color="#ff0000")

# Menu inicial
def sgMenu():
    layoutMenu = [
        [sg.Text("CENTRAL DE FERRAMENTARIA", font="Roboto 30", text_color="#2f56d8")],
        [sg.HSeparator()],
        [sg.Text("")],
        [sg.Button("CADASTRAR", size=(20))],
        [sg.Button("LISTAR", size=(20))],
        [sg.Button("EDITAR", size=(20))],
        [sg.Text("")],
        [sg.Button("Sair", button_color="#6aa84f", font="Roboto 15", size=(15))],
    ]
    
    janelaMenu = sg.Window("Central de ferramentaria", layout=layoutMenu, size=(500,300), font="Roboto 15", element_justification="center")
    
    while True:
        event, values = janelaMenu.read()
        if event == "CADASTRAR":
            janelaMenu.close()
            sgMenuCadastrar()
        if event == "LISTAR":
            janelaMenu.close()
            sgListar()
        if event == "EDITAR":
            janelaMenu.close()
            sgEditar()
        if event == "Sair":
            janelaMenu.close()
            sg.Exit()
            break
        elif event == sg.WIN_CLOSED:
            break

# Menu de cadastros
def sgMenuCadastrar():
    layoutMenuCadastrar = [
        [sg.Text("CADASTRAR", font="Roboto 30", text_color="#2f56d8")],
        [sg.HSeparator()],
        [sg.Text()],
        [sg.Button("FERRAMENTAS", size=(20)),sg.Button("TÉCNICOS", size=(20))],
        [sg.Button("RESERVAS", size=(20))],
        [sg.Text()],
        [sg.Button("Voltar", button_color="#6aa84f")],
        ]
        
    janelaMenuCadastrar = sg.Window("Central de Ferramentaria - Cadastrar", layout=layoutMenuCadastrar, size=(500,300), element_justification="center", font="Roboto 15")

    while True:
        event, values = janelaMenuCadastrar.read()
        if event =="FERRAMENTAS":
            janelaMenuCadastrar.close()
            sgCadFerramentas()
        elif event =="TÉCNICOS":
            janelaMenuCadastrar.close()
            sgCadTecnicos()
        elif event =="RESERVAS":
            janelaMenuCadastrar.close()
            sgCadReservas()
        elif event =="Voltar":
            janelaMenuCadastrar.close()
            sgMenu()
        elif event == sg.WIN_CLOSED:
            break

# Cadastro de Ferramentas
def sgCadFerramentas():   
    layoutCadFerramentas = [ 
                [sg.Text("Ferramenta"), sg.InputText("", key="FERRAMENTA")],
                [sg.Text("Fabricante"), sg.InputText("", key="FABRICANTE")],
                [sg.Text("Part Number"), sg.InputText("", key="PARTNUMBER")],
                [sg.Text("Tamanho"), sg.InputText("", key="TAMANHO")],
                [sg.Text("Medida"), sg.InputText("", key="MEDIDA")],
                [sg.Text("Material"), sg.InputText("", key="MATERIAL")],
                [sg.Text("Voltagem"), sg.InputText("", key="VOLTAGEM")],
                [sg.Text("Tipo"), sg.InputText("", key="TIPO")],
                [sg.Submit("Cadastrar", size=(10)), sg.Button("Voltar", button_color="#6aa84f", size=(10))] 
    ]

    janelaCadFerramentas = sg.Window("Cadastro de Ferramentas", layout=layoutCadFerramentas, font="Roboto 15", element_justification="center")
    
    while True:
        event, values = janelaCadFerramentas.read()
        if event == "Cadastrar":
            sg.Popup("Ferramenta cadastrada com sucesso!", font="Roboto 20", 
            background_color="#2f56d8", 
            auto_close=True, 
            auto_close_duration=10,
            no_titlebar=True)
            lista = list(values.values())
            adicionaInfo("Ferramentas.xlsx", lista)
            janelaCadFerramentas.close()
            sgCadFerramentas()
        if event == "Voltar":
            janelaCadFerramentas.close()
            sgMenuCadastrar()
        if event == sg.WIN_CLOSED:
            sg.Exit()
            break

# Cadastro de técnicos
def sgCadTecnicos():
    layoutCadTecnicos = [ 
            [sg.Text("CPF"), sg.InputText(key="CPF")],
            [sg.Text("Nome"), sg.InputText(key="Nome")],
            [sg.Text("Sobrenome"), sg.InputText(key="Sobrenome")],
            [sg.Text("Telefone"), sg.InputText(key="Telefone")],
            [sg.Text("Turno: "), sg.Radio("Manhã", "RADIO", key="Manha"), sg.Radio("Tarde", "RADIO", key="Tarde"), sg.Radio("Noite", "RADIO", key="Noite")],
            [sg.Text("Equipe"), sg.InputText(key="Equipe")],
            [sg.Submit("Cadastrar", size=(10)), sg.Button("Voltar", button_color="#6aa84f", size=(10))] 
            ]

    janelaCadTecnicos = sg.Window("CADASTRO DE TÉCNICOS", layout=layoutCadTecnicos, font="Roboto 15", element_justification="center")
    
    while True:
        event, values = janelaCadTecnicos.read()
        if event == "Cadastrar":
            sg.Popup("Técnico Cadastrado", font="Roboto 20", 
            background_color="#2f56d8", 
            auto_close=True, 
            auto_close_duration=10,
            no_titlebar=True)
            lista = list(values.values())
            adicionaInfo("Tecnicos.xlsx", lista)
            janelaCadTecnicos.close()
            sgCadTecnicos()
        if event == "Voltar":
            janelaCadTecnicos.close()
            sgMenuCadastrar()
        if event == sg.WIN_CLOSED:
            sg.Exit()
            break

# Cadastro de Reservas
def sgCadReservas():
    layoutCadR = [ 
            [sg.Text("Tecnico"), sg.InputText(key="TECNICO")],
            [sg.Text("Ferramenta"), sg.InputText(key="FERRAMENTA")],
            [sg.Input(key="DT-RESERVA", size=(20,1)) , sg.CalendarButton("Data da Reserva", close_when_date_chosen=True, target="DT-RESERVA",location=(800,400),no_titlebar=False)],
            [sg.Input(key="DT-DEVOLUCAO", size=(20,1)), sg.CalendarButton("Data da Devolução", close_when_date_chosen=True, target="DT-DEVOLUCAO",location=(800,400),no_titlebar=False)],
            [sg.Submit("Cadastrar", size=(10)), sg.Button("Voltar", button_color="#6aa84f", size=(10))] 
            ]

    janelaCadReservas = sg.Window("Cadastro de Reserva", layout=layoutCadR, font="Roboto 15", element_justification="center")
    
    while True:
        event, values = janelaCadReservas.read()
        if event == "Cadastrar":
            sg.Popup("Ferramenta reservada com sucesso!", font="Roboto 20", 
            background_color="#2f56d8", 
            auto_close=True, 
            auto_close_duration=10,
            no_titlebar=True)
            listaTemp = list(values.values())
            lista=[]
            for item in listaTemp:
                if item not in ["", None]:
                    lista.append(item)
            adicionaInfo("Reservas.xlsx", lista)
            janelaCadReservas.close()
            sgCadReservas()
        if event == "Voltar":
            janelaCadReservas.close()
            sgMenuCadastrar()
        if event == sg.WIN_CLOSED:
            sg.Exit()
            break

#Menu Listar
def sgListar():
    layoutMenuCadastrar = [
        [sg.Text("LISTAR", font="Roboto 30", text_color="#2f56d8")],
        [sg.HSeparator()],
        [sg.Text()],
        [sg.Button("FERRAMENTAS", size=(20)), sg.Button("TÉCNICOS", size=(20))],
        [sg.Button("RESERVAS", size=(20))],
        [sg.Text()],
        [sg.Button("Voltar", button_color="#6aa84f", size=(20))],
        ]
        
    janelaListar = sg.Window("Menu Listar", layout=layoutMenuCadastrar, size=(500,300), element_justification="center", font="Roboto 15")

    while True:
        event, values = janelaListar.read()
        if event =="FERRAMENTAS":
            janelaListar.close()
            sgConsultaF()
        elif event =="TÉCNICOS":
            janelaListar.close()
            sgConsultaT()
        elif event =="RESERVAS":
            janelaListar.close()
            sgConsultaR()
        elif event =="Voltar":
            janelaListar.close()
            sgMenu()
        elif event == sg.WIN_CLOSED:
            break

# Consulta de Ferramentas
def sgConsultaF(lista=False, values="", colunas=[1,2,3]):
    if lista == False:
        listaF=[]
    else:
        listaF = CarregarTabela("Ferramentas.xlsx", filtro=True, values = values, colunas = colunas)
    
    listaF_head = ["Ferramentas","Fabricante","Part Number","Material","Tamanho","Medida","Material","Voltagem","Tipo"]
    layoutConsultaF_topo = [ 
            [sg.Text("Ferramenta"), sg.InputText(key="-CS-Ferramenta-")],
            [sg.Text("Fabricante"), sg.InputText(key="-CS-Fabricante-")],
            [sg.Text("Part Number"), sg.InputText(key="-CS-Part-Number-")],
            [sg.Submit("Consultar", size=(10)), sg.Button("Voltar", button_color="#6aa84f", size=(10))],
            ]
    
    layoutConsultaF_baixo = [
        [sg.Table(values=listaF, headings=listaF_head ,max_col_width=35,
        auto_size_columns=True,
        justification="center",
        num_rows=10,
        key="-TABELA-",
        row_height=35)],
    ]

    layoutConsultaF = [
        [sg.Column(layoutConsultaF_topo)],
        [sg.HSeparator()],
        [sg.Column(layoutConsultaF_baixo)],

    ]

    janelaConsultaF = sg.Window("Consulta de Ferramentas", layout=layoutConsultaF, font="Roboto 15", element_justification="center")

    while True:
        event, values = janelaConsultaF.read()
        if event == "Consultar":
            lista = list(values.values())
            janelaConsultaF.close()
            sgConsultaF(lista=True, values=lista)
        if event == "Voltar":
            janelaConsultaF.close()
            sgListar()
            break
        elif event == sg.WIN_CLOSED:
            break

# Consulta de Técnicos
def sgConsultaT(lista=False, values="", colunas=[1,2,3]):
    if lista == False:
        listaT=[]
    else:
        listaT = CarregarTabela("Tecnicos.xlsx", filtro=True, values = values, colunas = colunas)
    
    listaT_head = ["CPF","Nome","Sobrenome","Telefone","Turno"]
    layoutConsultaT_topo = [ 
            [sg.Text("CPF"), sg.InputText(key="-CPF-")],
            [sg.Text("Nome"), sg.InputText(key="-NOME-")],
            [sg.Text("Sobrenome"), sg.InputText(key="-SOBRENOME-")],
            [sg.Submit("Consultar", size=(10)), sg.Button("Voltar", button_color="#6aa84f", size=(10))],
            ]
    
    layoutConsultaT_baixo = [
        [sg.Table(values=listaT, headings=listaT_head ,max_col_width=35,
        auto_size_columns=True,
        justification="center",
        num_rows=10,
        key="-TABELA-",
        row_height=35)],
    ]

    layoutConsultaT = [
        [sg.Column(layoutConsultaT_topo)],
        [sg.HSeparator()],
        [sg.Column(layoutConsultaT_baixo)],
    ]

    janelaConsultaT = sg.Window("Consulta de Técnicos", layout=layoutConsultaT, font="Roboto 15", element_justification="center")

    while True:
        event, values = janelaConsultaT.read()
        if event == "Consultar":
            lista = list(values.values())
            janelaConsultaT.close()
            sgConsultaT(lista=True, values=lista)
        if event == "Voltar":
            janelaConsultaT.close()
            sgListar()
            break
        elif event == sg.WIN_CLOSED:
            break

# Consulta de Reservas
def sgConsultaR(lista=False, values="", colunas=[1,2,3]):
    if lista == False:
        listaR=[]
    else:
        listaR = CarregarTabela("Reservas.xlsx", filtro=True, values = values, colunas = colunas)
    
    listaR_head = ["Técnico","Ferramenta","ReservaS","Devolução"]
    layoutConsultaR_topo = [ 
            [sg.Text("Técnico"), sg.InputText(key="-Tecnico-")],
            [sg.Text("Ferramenta"), sg.InputText(key="-Ferramenta-")],
            [sg.Input(key="DT-RESERVA", size=(20,1)) , sg.CalendarButton("Data da Reserva", close_when_date_chosen=True, target="DT-RESERVA",location=(800,400),no_titlebar=False)],
            [sg.Submit("Consultar", size=(10)), sg.Button("Voltar", button_color="#6aa84f", size=(10))],
            ]
    
    layoutConsultaR_baixo = [
        [sg.Table(values=listaR, headings=listaR_head ,max_col_width=35,
        auto_size_columns=True,
        justification="center",
        num_rows=10,
        key="-TABELA-",
        row_height=35)],
    ]

    layoutConsultaR = [
        [sg.Column(layoutConsultaR_topo)],
        [sg.HSeparator()],
        [sg.Column(layoutConsultaR_baixo)],
    ]

    janelaConsultaR = sg.Window("Consulta de Reservas", layout=layoutConsultaR, font="Roboto 15", element_justification="center")

    while True:
        event, values = janelaConsultaR.read()
        if event == "Consultar":
            lista = list(values.values())
            janelaConsultaR.close()
            sgConsultaR(lista=True, values=lista)
        if event == "Voltar":
            janelaConsultaR.close()
            sgListar()
            break
        elif event == sg.WIN_CLOSED:
            break

# Menu Editar
def sgEditar():
    layoutMenuCadastrar = [
        [sg.Text("EDITAR", font="Roboto 30", text_color="#2f56d8")],
        [sg.HSeparator()],
        [sg.Text()],
        [sg.Button("FERRAMENTAS", size=(20)), sg.Button("TÉCNICOS", size=(20))],
        [sg.Button("RESERVAS", size=(20))],
        [sg.Text()],
        [sg.Button("Voltar", button_color="#6aa84f")],
        ]
        
    janelaMenuCadastrar = sg.Window("EDITAR", layout=layoutMenuCadastrar, size=(500,300), element_justification="center", font="Roboto 15")

    while True:
        event, values = janelaMenuCadastrar.read()
        if event =="FERRAMENTAS":
            janelaMenuCadastrar.close()
            sgEditar()
        elif event =="TÉCNICOS":
            janelaMenuCadastrar.close()
            sgEditar()
        elif event =="RESERVAS":
            janelaMenuCadastrar.close()
            sgEditar()
        elif event =="Voltar":
            janelaMenuCadastrar.close()
            sgMenu()
        elif event == sg.WIN_CLOSED:
            break

# Gerenciamento de dados nas planilhas
def AbrirBD(caminho):
    bd = opx.load_workbook(caminho)
    return bd

def FecharBD(caminho, bd):
    bd.save(caminho)
    bd.close()

def adicionaInfo(caminho, lista):
    wb = AbrirBD(caminho)
    ws = wb["Sheet1"]
    ultimaL = 1
    for lin in range(1,1000000): # verificação das linhas
        if ws.cell(row=lin, column=1).value == None:
            ultimaL = lin
            break
    for col in range(1, len(lista)+1): # Add valores na coluna
        ws.cell(row=ultimaL, column=col).value = lista[col-1]

    FecharBD(caminho, wb)

# Filtros de pesquisa
def CarregarTabela(caminho, values="", filtro=False, colunas=""):
    wb = AbrirBD(caminho)
    ws = wb["Sheet1"]
    tabela = []
    linha = []
    checkFiltro = True
    for lin in range(2,1000000):
        if ws.cell(row=lin, column=1).value == None: break
        linha = []
        checkFiltro = True
        if filtro == True:
            cont = 0
            for col in colunas:
                if values[cont] in ["", []]:
                    cont += 1
                    continue
                if str(ws.cell(row=lin, column=col).value).find(values[cont]) == -1:  # verificando coluna via filtro
                    checkFiltro = False
                    break
                cont += 1
        if checkFiltro == False : continue

        for col in range(1,1000):
            if ws.cell(row=lin, column=col).value == None: break
            linha.append(ws.cell(row=lin, column=col).value)
        tabela.append(linha)

    FecharBD(caminho, wb)
    return tabela

sg.theme("Black")
sgLogin()